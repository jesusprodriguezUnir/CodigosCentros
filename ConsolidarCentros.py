import pandas as pd
import os
from functools import reduce
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

BASE = r'C:\Temp\Noemi\Centros'
REPO = os.path.dirname(os.path.abspath(__file__))

PRIORIDAD = ['2526_rh09', '2526_anexo_i', '2526_anexo_via', '2425', '2324', '2223']

FUENTES = [
    {
        'nombre': '2223',
        'path': os.path.join(REPO, 'agrupado_vacantes2223.xlsx'),
        'col_cod': 'Codigo',
        'col_loc': 'Localidad',
        'col_centro': 'Nombre del centro',
        'col_vac': 'Total Vacantes',
        'skip_header_row': False,
    },
    {
        'nombre': '2324',
        'path': os.path.join(REPO, 'agrupado_vacantes.xlsx'),
        'col_cod': 'Codigo',
        'col_loc': 'Localidad',
        'col_centro': 'Nombre del centro',
        'col_vac': 'Total Vacantes',
        'skip_header_row': True,
    },
    {
        'nombre': '2425',
        'path': os.path.join(REPO, 'agrupado_columna_2425.xlsx'),
        'col_cod': 'Codigo',
        'col_loc': 'LOCALIDAD',
        'col_centro': 'CENTRO',
        'col_vac': 'Total Vacantes',
        'skip_header_row': False,
    },
    {
        'nombre': '2526_rh09',
        'path': os.path.join(REPO, '2026', 'salida', 'agrupado_2526.xlsx'),
        'col_cod': 'Cod_Centro',
        'col_loc': 'Localidad',
        'col_centro': 'Centro',
        'col_vac': 'Total Vacantes',
        'skip_header_row': False,
    },
    {
        'nombre': '2526_anexo_i',
        'path': os.path.join(REPO, '2026', 'salida', 'agrupado_anexo_i_2526.xlsx'),
        'col_cod': 'Cod_Centro',
        'col_loc': 'Localidad',
        'col_centro': 'Centro',
        'col_cp': 'CP',
        'col_vac': 'Total Vacantes',
        'skip_header_row': False,
    },
    {
        'nombre': '2526_anexo_via',
        'path': os.path.join(REPO, '2026', 'salida', 'agrupado_anexo_via_2526.xlsx'),
        'col_cod': 'Cod_Centro',
        'col_loc': 'Localidad',
        'col_centro': 'Centro',
        'col_cp': 'CP',
        'col_vac': 'Total Vacantes',
        'skip_header_row': False,
    },
]

OUT_DIR = os.path.join(REPO, '2026', 'salida')
OUT_PATH = os.path.join(OUT_DIR, 'consolidado_2526.xlsx')
JSON_DIR = os.path.join(REPO, 'web', 'datos')
JSON_PATH = os.path.join(JSON_DIR, 'consolidado.json')


def normalizar_codigo(v):
    try:
        s = str(int(float(v)))
        return s.zfill(8)
    except (ValueError, TypeError):
        return None


def normalizar_cp(v):
    if pd.isna(v):
        return ''
    try:
        s = str(int(float(v)))
        return s.zfill(5)
    except (ValueError, TypeError):
        return ''


def normalizar_texto(v):
    if pd.isna(v):
        return ''
    return ' '.join(str(v).replace('\n', ' ').split())


def cargar_fuente(spec):
    path = spec['path']
    if not os.path.exists(path):
        print(f'  [AVISO] No encontrado: {path}')
        return None

    df = pd.read_excel(path).copy()

    if spec['skip_header_row']:
        mask = df[spec['col_cod']].astype(str).str.strip().str.lower().isin(['c\u00f3digo', 'codigo'])
        if mask.any():
            df = df[~mask].copy()

    codigos = df[spec['col_cod']].apply(normalizar_codigo)
    vacantes = pd.to_numeric(df[spec['col_vac']], errors='coerce').fillna(0).astype(int)
    localidades = df[spec['col_loc']].apply(normalizar_texto)
    centros = df[spec['col_centro']].apply(normalizar_texto)

    cps = df[spec['col_cp']].apply(normalizar_cp) if 'col_cp' in spec else ''

    result = pd.DataFrame({
        'Cod_Centro': codigos,
        'Localidad': localidades,
        'Centro': centros,
        'Vacantes': vacantes,
    })
    if 'col_cp' in spec:
        result['CP'] = cps

    invalid = result['Cod_Centro'].isna()
    if invalid.any():
        result = result[~invalid].copy()

    agg_dict = {
        'Localidad': 'first',
        'Centro': 'first',
        'Vacantes': 'sum',
    }
    if 'col_cp' in spec:
        agg_dict['CP'] = 'first'

    result = result.groupby('Cod_Centro', as_index=False).agg(agg_dict)

    nombre = spec['nombre']
    rename_dict = {
        'Vacantes': f'Vacantes_{nombre}',
        'Localidad': f'Localidad_{nombre}',
        'Centro': f'Centro_{nombre}',
    }
    if 'col_cp' in spec:
        rename_dict['CP'] = f'CP_{nombre}'
    result.rename(columns=rename_dict, inplace=True)
    return result


def construir_consolidado():
    print('Cargando fuentes...')
    dfs = []
    stats = {}
    for f in FUENTES:
        print(f'  {f["nombre"]}: {f["path"]}')
        df = cargar_fuente(f)
        if df is not None:
            dfs.append(df)
            stats[f['nombre']] = len(df)
            print(f'    -> {len(df)} centros')

    print(f'\nTotal fuentes cargadas: {len(dfs)}')
    if len(dfs) < 6:
        print('  [AVISO] No se cargaron todas las 6 fuentes')

    merged = reduce(
        lambda left, right: pd.merge(left, right, on='Cod_Centro', how='outer'),
        dfs
    )

    merged['Localidad'] = pd.NA
    merged['Centro'] = pd.NA
    merged['CP'] = pd.NA
    for nombre in PRIORIDAD:
        loc_col = f'Localidad_{nombre}'
        cen_col = f'Centro_{nombre}'
        cp_col = f'CP_{nombre}'
        if loc_col in merged.columns:
            merged[loc_col] = merged[loc_col].replace('', pd.NA)
            merged['Localidad'] = merged['Localidad'].fillna(merged[loc_col])
        if cen_col in merged.columns:
            merged[cen_col] = merged[cen_col].replace('', pd.NA)
            merged['Centro'] = merged['Centro'].fillna(merged[cen_col])
        if cp_col in merged.columns:
            merged[cp_col] = merged[cp_col].replace('', pd.NA)
            merged['CP'] = merged['CP'].fillna(merged[cp_col])

    suf_cols = [c for c in merged.columns if c.startswith('Localidad_') or c.startswith('Centro_') or c.startswith('CP_')]
    merged.drop(columns=suf_cols, inplace=True)

    vac_cols = [c for c in merged.columns if c.startswith('Vacantes_')]
    for col in vac_cols:
        merged[col] = pd.to_numeric(merged[col], errors='coerce').fillna(0).astype(int)

    base_cols = ['Vacantes_2223', 'Vacantes_2324', 'Vacantes_2425',
                 'Vacantes_2526_rh09', 'Vacantes_2526_anexo_i', 'Vacantes_2526_anexo_via']
    merged['Total'] = sum(merged.get(c, 0) for c in base_cols)

    merged['Centro'] = merged['Centro'].fillna('')
    merged['Localidad'] = merged['Localidad'].fillna('')
    merged['CP'] = merged['CP'].fillna('')

    column_order = [
        'Cod_Centro', 'Centro', 'Localidad', 'CP',
        'Vacantes_2223', 'Vacantes_2324', 'Vacantes_2425',
        'Vacantes_2526_rh09', 'Vacantes_2526_anexo_i', 'Vacantes_2526_anexo_via',
        'Total',
    ]
    existing_cols = [c for c in column_order if c in merged.columns]
    merged = merged[existing_cols]

    merged.sort_values('Cod_Centro', inplace=True)
    merged.reset_index(drop=True, inplace=True)

    os.makedirs(OUT_DIR, exist_ok=True)
    actual_out = OUT_PATH
    try:
        with open(actual_out, 'ab'):
            pass
    except PermissionError:
        actual_out = os.path.join(OUT_DIR, 'consolidado_2526_nuevo.xlsx')
        print(f'[AVISO] {OUT_PATH} está abierto en Excel. Guardando en {actual_out}')

    merged.to_excel(actual_out, index=False, sheet_name='Consolidado')
    wb = load_workbook(actual_out)
    ws = wb['Consolidado']
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    headers = {cell.value: cell.column for cell in ws[1]}
    last_row = ws.max_row
    if last_row > 1:
        for col_name in base_cols:
            if col_name in headers:
                col_letter = get_column_letter(headers[col_name])
                rng = f'{col_letter}2:{col_letter}{last_row}'
                ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['10'], fill=red_fill))
        if 'Total' in headers:
            col_letter = get_column_letter(headers['Total'])
            rng = f'{col_letter}2:{col_letter}{last_row}'
            ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['40'], fill=red_fill))

    wb.save(actual_out)
    wb.close()
    print(f'\nConsolidado guardado en: {actual_out}')
    print(f'  Total filas: {len(merged)}')

    os.makedirs(JSON_DIR, exist_ok=True)
    merged.to_json(JSON_PATH, orient='records', force_ascii=False)
    print(f'JSON guardado en: {JSON_PATH}')

    print('\n=== Verificacion ===')
    for nombre, count in sorted(stats.items()):
        print(f'  {nombre}: {count} centros unicos tras limpieza')
    print(f'  Consolidado: {len(merged)} filas')

    testigo = merged[merged['Cod_Centro'] == '28071802']
    if not testigo.empty:
        print(f'\n  Centro testigo 28071802:')
        for c in testigo.columns:
            print(f'    {c}: {testigo.iloc[0][c]}')
    else:
        print('\n  Centro testigo 28071802 NO encontrado en consolidado')

    null_cod = merged['Cod_Centro'].isna().sum()
    null_centro = merged['Centro'].isna().sum()
    print(f'\n  NaN en Cod_Centro: {null_cod}')
    print(f'  NaN en Centro: {null_centro}')

    for col in base_cols:
        if col in merged.columns:
            dtype_ok = merged[col].dtype in ('int64', 'int32', 'int')
            assert dtype_ok, f'{col} no es int (es {merged[col].dtype})'
            assert (merged[col] >= 0).all(), f'{col} tiene valores negativos'
    print('  Todas las columnas de vacantes son int >= 0: OK')


if __name__ == '__main__':
    construir_consolidado()
